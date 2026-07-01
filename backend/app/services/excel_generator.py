import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any

# Define a nice palette
HEADER_FILL = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid") # Dark Blue
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

# Borders
THIN_BORDER = Border(
    left=Side(style='thin', color='CBD5E0'),
    right=Side(style='thin', color='CBD5E0'),
    top=Side(style='thin', color='CBD5E0'),
    bottom=Side(style='thin', color='CBD5E0')
)

# Text styling
DATA_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", size=10, bold=True)

# Income / Expense fills
INCOME_FILL = PatternFill(start_color="E6FFFA", end_color="E6FFFA", fill_type="solid") # Light Green/Teal
EXPENSE_FILL = PatternFill(start_color="FFF5F5", end_color="FFF5F5", fill_type="solid") # Light Red/Pink

# Status fills (optional highlights)
STATUS_COLORS = {
    "Новый": "EBF8FF",          # Light Blue
    "Переговоры": "FEFCBF",     # Light Yellow
    "Выезд на аудит": "E2E8F0",  # Light Gray
    "КП отправлено": "FAF089",   # Yellow
    "Договор": "C6F6D5",        # Green
    "В работе": "BEE3F8",       # Blue
    "Завершено": "E2E8F0",      # Gray
    "Анализ": "EDF2F7",
    "Участие": "EBF8FF",
    "Заявка подана": "FEFCBF",
    "Выигран": "C6F6D5",
    "Проигран": "FED7D7",
    "Отклонен": "E2E8F0"
}

def auto_fit_columns(ws, min_width=12, max_width=40):
    """
    Adjusts the column widths based on their content length.
    """
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            # Skip formulas or None
            if cell.value is not None:
                val_str = str(cell.value)
                # Ignore very long strings for length calculation (e.g. notes)
                if len(val_str) < 100:
                    max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(min(max_len + 3, max_width), min_width)

def create_styled_workbook() -> tuple[Workbook, Any]:
    wb = Workbook()
    ws = wb.active
    ws.views.sheetView[0].showGridLines = True
    return wb, ws

def generate_clients_excel(clients: List[Any]) -> bytes:
    wb, ws = create_styled_workbook()
    ws.title = "Клиенты"

    headers = [
        "ID", "Название компании", "ИНН", "КПП", "ОГРН", 
        "Сегмент", "Статус", "Контактное лицо", "Телефон", "Email", 
        "Банк", "БИК", "Р/С", "К/С", "Юридический адрес", 
        "Стоимость привлечения (₽)", "Дата создания", "Заметки"
    ]

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    
    ws.row_dimensions[1].height = 28

    # Write data
    for row_idx, client in enumerate(clients, 2):
        created_at_str = ""
        if client.created_at:
            created_at_str = client.created_at.strftime("%Y-%m-%d %H:%M:%S")

        values = [
            client.id,
            client.name or "",
            client.inn or "",
            client.kpp or "",
            client.ogrn or "",
            client.segment.value if client.segment else "",
            client.status.value if client.status else "",
            client.contact_person or "",
            client.phone or "",
            client.email or "",
            client.bank_name or "",
            client.bik or "",
            client.rs or "",
            client.ks or "",
            client.legal_address or "",
            client.acquisition_cost or 0.0,
            created_at_str,
            client.notes or ""
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            
            # Alignments & formatting
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in [3, 4, 5, 12]:  # INN, KPP, OGRN, BIK
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = "@" # Text format to keep leading zeros
            elif col_idx in [13, 14]:  # RS, KS
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = "@"
            elif col_idx == 7: # Status
                cell.alignment = Alignment(horizontal="center")
                status_val = str(val)
                if status_val in STATUS_COLORS:
                    cell.fill = PatternFill(start_color=STATUS_COLORS[status_val], end_color=STATUS_COLORS[status_val], fill_type="solid")
            elif col_idx == 16:  # Acquisition Cost
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'
            elif col_idx == 17:  # Created At
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

    auto_fit_columns(ws)
    
    # Save to buffer
    fp = io.BytesIO()
    wb.save(fp)
    return fp.getvalue()

def generate_finance_excel(transactions: List[Any]) -> bytes:
    wb, ws = create_styled_workbook()
    ws.title = "Финансы"

    headers = [
        "ID", "Дата", "Тип", "Сумма (₽)", "Касса", 
        "Способ оплаты", "Категория", "Клиент", "Объект", "Описание"
    ]

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 28

    # Write data
    for row_idx, trans in enumerate(transactions, 2):
        date_str = ""
        if trans.date:
            date_str = trans.date.strftime("%Y-%m-%d %H:%M:%S")

        tr_type_rus = "Поступление" if trans.transaction_type == "income" else "Расход"
        cash_register_rus = "Работы" if trans.cash_register == "works" else "Материалы"

        values = [
            trans.id,
            date_str,
            tr_type_rus,
            trans.amount or 0.0,
            cash_register_rus,
            trans.payment_method or "",
            trans.category or "",
            trans.client_name or "",
            trans.object_name or "",
            trans.description or ""
        ]

        # Determine row background or transaction type background
        row_fill = INCOME_FILL if trans.transaction_type == "income" else EXPENSE_FILL

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            cell.fill = row_fill

            # Alignments & formatting
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 2:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 3:
                cell.alignment = Alignment(horizontal="center")
                cell.font = BOLD_FONT
            elif col_idx == 4:
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'
                cell.font = BOLD_FONT
            elif col_idx in [5, 6, 7]:
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

    auto_fit_columns(ws)

    # Save to buffer
    fp = io.BytesIO()
    wb.save(fp)
    return fp.getvalue()

def generate_tenders_excel(tenders: List[Any]) -> bytes:
    wb, ws = create_styled_workbook()
    ws.title = "Тендеры"

    headers = [
        "ID", "Номер тендера", "Название тендера", "Заказчик", "ИНН", 
        "Цена", "Валюта", "Площадка", "Ссылка", "Статус", 
        "Дата публикации", "Дедлайн подачи", "Ответственный сотрудник", 
        "Клиент", "Объект", "Описание", "Дата создания"
    ]

    # Write headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    ws.row_dimensions[1].height = 28

    # Write data
    for row_idx, tender in enumerate(tenders, 2):
        created_at_str = tender.created_at.strftime("%Y-%m-%d %H:%M:%S") if tender.created_at else ""
        pub_date_str = tender.publication_date.strftime("%Y-%m-%d %H:%M:%S") if tender.publication_date else ""
        deadline_str = tender.submission_deadline.strftime("%Y-%m-%d %H:%M:%S") if tender.submission_deadline else ""

        values = [
            tender.id,
            tender.tender_number or "",
            tender.title or "",
            tender.customer_name or "",
            tender.inn or "",
            tender.price or 0.0,
            tender.currency or "RUB",
            tender.platform or "",
            tender.link or "",
            tender.status or "",
            pub_date_str,
            deadline_str,
            tender.assigned_username or "",
            tender.client_name or "",
            tender.object_name or "",
            tender.description or "",
            created_at_str
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER

            # Alignments & formatting
            if col_idx == 1:
                cell.alignment = Alignment(horizontal="center")
            elif col_idx in [2, 5]:  # Tender number, INN
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = "@"
            elif col_idx == 6:  # Price
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = '#,##0.00'
            elif col_idx == 7:  # Currency
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 9 and val:  # Link
                cell.value = "Перейти"
                cell.hyperlink = val
                cell.font = Font(name="Calibri", size=10, color="0000FF", underline="single")
                cell.alignment = Alignment(horizontal="center")
            elif col_idx == 10:  # Status
                cell.alignment = Alignment(horizontal="center")
                status_val = str(val)
                if status_val in STATUS_COLORS:
                    cell.fill = PatternFill(start_color=STATUS_COLORS[status_val], end_color=STATUS_COLORS[status_val], fill_type="solid")
            elif col_idx in [11, 12, 17]:  # Dates
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left")

    auto_fit_columns(ws)

    # Save to buffer
    fp = io.BytesIO()
    wb.save(fp)
    return fp.getvalue()
